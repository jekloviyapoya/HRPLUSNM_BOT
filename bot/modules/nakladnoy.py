"""Nakladnoy moduli: hujjatni o'qish va tekshirish.

Bosqichlar: rasm/PDF/Excel yuborish → AI o'qiydi → foydalanuvchi
tekshiradi → (keyingi bosqichda) Bito'ga kirim.

Bu fayl Bito'ga HECH NARSA YOZMAYDI. Yozish alohida bosqichda, moslashtirish
tayyor bo'lgandan keyin qo'shiladi.

────────────────────────────────────────────────────────────────────
INVARIANT — buzilmasin
────────────────────────────────────────────────────────────────────
    qty         = BLOK soni
    block_size  = 1 blokdagi dona
    price       = BITTA DONA narxi
    jami        = qty × block_size × price

market-bot'da (2026-08-02) qty donaga aylantirilgan, keyin boshqa joy
YANA ko'paytirgan: 135 000 o'rniga 810 000 chiqib, Bito'ga olti barobar
ortiq yuklangan. Shuning uchun bu yerda qty hech qachon donaga
aylantirilmaydi — faqat block_size aniqlanadi.
"""

import io
import itertools
import logging

from . import base, nak_prompt, registry
from .. import (ai, bito, catalog, ctx, db, nak_upload, sessions,
                tenant, ui, users)
from ..errors import BotError

log = logging.getLogger(__name__)

MAX_FILE_MB = 12
BLOCK_MIN, BLOCK_MAX = 2, 96
BLOCK_TOLERANCE = 0.02


# ---------------------------------------------------------------- normallash


def clean_barcode(value):
    """Faqat raqamlar, 8–14 xonali. AI ba'zan artikulni ilashtiradi."""
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    return digits if 8 <= len(digits) <= 14 else ""


def _number(value, default=0.0):
    if value in (None, ""):
        return default
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return default


def detect_block(qty, price, doc_total):
    """Hujjatdagi qator jamisidan blok hajmini teskari hisoblaydi.

    qty × price hujjat jamisiga to'g'ri kelmasa va nisbat butun songa
    yaqin bo'lsa — demak qty blokda, price esa donada berilgan.

    Faqat 2..96 oralig'idagi butun nisbat qabul qilinadi: aks holda
    tasodifiy yaxlitlash xatosi blok deb qabul qilinib ketadi.
    """
    if qty <= 0 or price <= 0 or not doc_total or doc_total <= 0:
        return 1
    expected = qty * price
    if abs(expected - doc_total) < 0.01:
        return 1
    ratio = doc_total / expected
    if ratio <= 1:
        return 1
    nearest = round(ratio)
    if abs(ratio - nearest) < BLOCK_TOLERANCE and BLOCK_MIN <= nearest <= BLOCK_MAX:
        return int(nearest)
    return 1


def line_total(item):
    """Yagona joy — hamma yerda shu ishlatiladi."""
    return item["qty"] * item["block_size"] * item["price"]


def normalize(raw):
    """AI javobini qatorlarga aylantiradi. Invariant shu yerda o'rnatiladi."""
    items = []
    for position, row in enumerate(raw.get("items") or []):
        name = str(row.get("name") or "").strip()
        if not name:
            continue
        qty = _number(row.get("qty"))
        price = _number(row.get("price"))
        doc_total = row.get("total")
        doc_total = _number(doc_total, default=None) if doc_total is not None \
            else None
        if qty <= 0:
            continue
        items.append({
            "position": position,
            "raw_name": name,
            "qty": qty,                       # BLOK soni — o'zgartirilmaydi
            "qty_unit": str(row.get("qty_unit") or "").strip() or None,
            "block_size": detect_block(qty, price, doc_total),
            "price": price,                   # BITTA DONA narxi
            "doc_total": doc_total,
            "barcode": clean_barcode(row.get("barcode")),
        })
    return {
        "supplier": str(raw.get("supplier") or "").strip(),
        "number": str(raw.get("number") or "").strip(),
        "date": str(raw.get("date") or "").strip(),
        "total": _number(raw.get("total"), default=None)
        if raw.get("total") is not None else None,
        "items": items,
    }


def check_totals(doc):
    """Hisoblangan jami hujjatdagi jamiga to'g'ri keladimi?

    Qaytadi: (hisoblangan, hujjatdagi, farq_foizda) yoki None.
    """
    computed = sum(line_total(item) for item in doc["items"])
    stated = doc.get("total")
    if not stated:
        return computed, None, None
    diff = abs(computed - stated) / stated * 100 if stated else 0
    return computed, stated, diff


# ------------------------------------------------------------------- saqlash


def create_doc(tg_id, source, file_id=None):
    cur = db.run(
        "INSERT INTO nak_doc (tenant_id, tg_id, source, file_id) "
        "VALUES (?, ?, ?, ?)",
        (ctx.require(), tg_id, source, file_id),
    )
    return cur.lastrowid


def save_doc(doc_id, parsed):
    db.run(
        "UPDATE nak_doc SET status = 'tekshirilmoqda', supplier_name = ?, "
        "  doc_number = ?, doc_date = ?, doc_total = ?, error = NULL, "
        "  updated_at = datetime('now') WHERE tenant_id = ? AND id = ?",
        (parsed["supplier"] or None, parsed["number"] or None,
         parsed["date"] or None, parsed["total"], ctx.require(), doc_id),
    )
    db.run("DELETE FROM nak_item WHERE tenant_id = ? AND doc_id = ?",
           (ctx.require(), doc_id))
    for item in parsed["items"]:
        db.run(
            "INSERT INTO nak_item (tenant_id, doc_id, position, raw_name, qty, "
            "  qty_unit, block_size, price, doc_total, barcode) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (ctx.require(), doc_id, item["position"], item["raw_name"],
             item["qty"], item["qty_unit"], item["block_size"], item["price"],
             item["doc_total"], item["barcode"]),
        )


def fail_doc(doc_id, message):
    db.run(
        "UPDATE nak_doc SET status = 'xato', error = ?, "
        "  updated_at = datetime('now') WHERE tenant_id = ? AND id = ?",
        (message[:300], ctx.require(), doc_id),
    )


def get_doc(doc_id):
    return db.row("SELECT * FROM nak_doc WHERE tenant_id = ? AND id = ?",
                  (ctx.require(), doc_id))


def get_items(doc_id):
    return db.rows(
        "SELECT * FROM nak_item WHERE tenant_id = ? AND doc_id = ? "
        "ORDER BY position",
        (ctx.require(), doc_id),
    )


def last_doc(tg_id):
    return db.row(
        "SELECT * FROM nak_doc WHERE tenant_id = ? AND tg_id = ? "
        "ORDER BY created_at DESC LIMIT 1",
        (ctx.require(), tg_id),
    )


def hints():
    return db.rows(
        "SELECT supplier, hint FROM nak_hint WHERE tenant_id = ? "
        "ORDER BY used_count DESC, updated_at DESC LIMIT 15",
        (ctx.require(),),
    )


def remember_hint(supplier, hint):
    if not supplier or not hint:
        return
    db.run(
        "INSERT INTO nak_hint (tenant_id, supplier, hint) VALUES (?, ?, ?) "
        "ON CONFLICT (tenant_id, supplier) DO UPDATE SET "
        "  hint = excluded.hint, used_count = used_count + 1, "
        "  updated_at = datetime('now')",
        (ctx.require(), supplier.strip().lower(), hint),
    )


# -------------------------------------------------------------------- o'qish


def excel_to_text(data, max_lines=300):
    """Excel faylni matnga — AI tahlili uchun."""
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(data), data_only=True)
    sheet = book.active
    lines = []
    for row in sheet.iter_rows(values_only=True):
        cells = [str(cell) for cell in row
                 if cell is not None and str(cell).strip()]
        if cells:
            lines.append(" | ".join(cells))
        if len(lines) >= max_lines:
            break
    return "\n".join(lines)


def extract(data=None, filename="", text=None, session=None):
    """Hujjatni o'qib, normallashtirilgan lug'at qaytaradi."""
    hint_rows = hints()

    if text is not None:
        content = [{"type": "text",
                    "text": nak_prompt.build(hint_rows, text=text)}]
    elif filename.lower().endswith(".pdf"):
        content = [ai.pdf_block(data),
                   {"type": "text", "text": nak_prompt.build(hint_rows)}]
    elif filename.lower().endswith((".xlsx", ".xls", ".xlsm")):
        as_text = excel_to_text(data)
        if not as_text.strip():
            raise BotError("Excel fayl bo'sh ko'rinadi.")
        content = [{"type": "text",
                    "text": nak_prompt.build(hint_rows, text=as_text)}]
    else:
        content = [ai.image_block(data, filename),
                   {"type": "text", "text": nak_prompt.build(hint_rows)}]

    raw, _ = ai.ask_json(content, session=session)
    parsed = normalize(raw)
    if not parsed["items"]:
        raise BotError(
            "Hujjatdan birorta mahsulot o'qilmadi. Rasm aniqroq bo'lsa yoki "
            "hujjat to'liq ko'rinsa qayta urining."
        )
    return parsed


# ------------------------------------------------------------ moslashtirish


def match_doc(doc_id):
    """Hamma qatorlarni katalogga moslashtiradi va saqlaydi."""
    items = get_items(doc_id)
    results = catalog.match_all([dict(row) for row in items])
    for row, found in zip(items, results):
        db.run(
            "UPDATE nak_item SET product_id = ?, product_name = ?, "
            "  match_state = ?, note = ? WHERE tenant_id = ? AND id = ?",
            (found["product_id"], found["product_name"],
             "topildi" if found["state"] == "topildi" else "yoq",
             found["how"], ctx.require(), row["id"]),
        )
    db.run("UPDATE nak_doc SET status = 'moslashtirilmoqda' "
           "WHERE tenant_id = ? AND id = ?", (ctx.require(), doc_id))
    return summary(doc_id)


def summary(doc_id):
    rows = get_items(doc_id)
    matched = [r for r in rows if r["match_state"] == "topildi"]
    skipped = [r for r in rows if r["match_state"] == "tashlab"]
    pending = [r for r in rows
               if r["match_state"] not in ("topildi", "tashlab")]
    return {"all": rows, "matched": matched, "pending": pending,
            "skipped": skipped}


def set_match(item_id, product_id, product_name):
    db.run(
        "UPDATE nak_item SET product_id = ?, product_name = ?, "
        "  match_state = 'topildi', note = 'qo''lda' "
        "WHERE tenant_id = ? AND id = ?",
        (product_id, product_name, ctx.require(), item_id),
    )
    row = db.row("SELECT raw_name FROM nak_item WHERE tenant_id = ? AND id = ?",
                 (ctx.require(), item_id))
    if row:
        catalog.remember(row["raw_name"], product_id, product_name)


def skip_item(item_id):
    db.run("UPDATE nak_item SET match_state = 'tashlab' "
           "WHERE tenant_id = ? AND id = ?", (ctx.require(), item_id))


# ------------------------------------------------- yangi mahsulot yaratish

_KG_HINTS = ("kg", "кг", "кг.", "килограмм", "kilogram")


def _looks_kg(name, qty_unit=None):
    text = f"{name or ''} {qty_unit or ''}".lower()
    return any(hint in text.split() or text.endswith(hint)
               for hint in _KG_HINTS)


_sku_seq = itertools.count()


def _new_sku():
    """Bito'da SKU majburiy va unikal. Do'kondagi raqamlar bilan
    to'qnashmasligi uchun 9 bilan boshlanadigan 8 xonali raqam: vaqt +
    hisoblagich (bir soniya ichida ham takrorlanmasin). To'qnashsa
    chaqiruvchi qayta uradi."""
    import time

    return f"9{(int(time.time() * 10) + next(_sku_seq)) % 10_000_000:07d}"


def create_in_bito(item_id, client=None):
    """Nakladnoy qatoridan Bito'da yangi mahsulot yaratadi.

    Jonli sinov saboqlari (2026-08-14, mijoz Bito'sida):
    - `organizations` kamida bitta obyekt bilan majburiy; qolgan
      maydonlarni (amount, standard...) server o'zi to'ldiradi
    - `barcode` topda yuboriladi, `barcodes[]` bo'sh qolsa ham skaner
      qidiruvi ishlaydi
    - PLU (custom_fields) yaratishda BERILMAYDI — tarozi kodini do'kon
      Bito'da o'zi belgilaydi; noto'g'ri PLU tarozida boshqa mahsulotni
      chiqarib yuboradi
    """
    row = db.row("SELECT * FROM nak_item WHERE tenant_id = ? AND id = ?",
                 (ctx.require(), item_id))
    if not row:
        raise BotError("Qator topilmadi.")
    organization = tenant.get("bito_org_id")
    if not organization:
        raise BotError("Bito tashkiloti tanlanmagan — Sozlamalarga kiring.")

    client = client or bito.client()
    uoms = client.uoms()
    if _looks_kg(row["raw_name"], row["qty_unit"]):
        uom = bito.pick_uom(uoms, "kilogram", ("kg", "Kilogram"))
    else:
        uom = bito.pick_uom(uoms, "piece", ("dona", "Dona", "sht", "шт"))
    if not uom:
        raise BotError("Bito'da o'lchov birligi topilmadi.")

    body = {
        "name": (row["raw_name"] or "").strip()[:120],
        "measure_id": str(uom["_id"] if isinstance(uom, dict) else uom),
        "organizations": [{
            "organization_id": str(organization),
            "is_available": True,
            "is_available_for_sale": True,
        }],
        "is_product": True,
        "is_material": False,
        "is_semi_product": False,
        "is_marked": False,
    }
    if row["barcode"]:
        body["barcode"] = str(row["barcode"])

    created = None
    for _attempt in range(2):          # SKU to'qnashsa bir marta qayta
        body["sku"] = _new_sku()
        try:
            created = bito.unwrap(client.create_product(body))
            break
        except bito.BitoError as e:
            if "sku" not in str(e).lower() or _attempt:
                raise
    pid = str((created or {}).get("_id") or "")
    if not pid:
        raise BotError("Bito mahsulotni qaytarmadi — qayta uring.")

    catalog.upsert(created)
    set_match(item_id, pid, created.get("name"))
    return created


def find_supplier(name, client=None):
    """Hujjatdagi firma nomiga Bito'dan mos keluvchini topadi."""
    if not name:
        return None, []
    client = client or bito.client()
    rows, _ = client.suppliers(page=1, limit=20, search=name.strip())
    if not rows:
        return None, []
    key = catalog.normalize(name)
    for row in rows:
        if catalog.normalize(row.get("name")) == key:
            return row, rows
    return None, rows[:5]


# -------------------------------------------------------------------- modul


@registry.implement("nakladnoy")
class Nakladnoy(base.Module):
    def menu(self, role):
        if role == "staff":
            return []
        return [("🧾 Nakladnoy", "mod:nakladnoy:panel")]

    def register(self, bot, guard):
        _register(bot, guard)


def _register(bot, guard):
    @bot.callback_query_handler(
        func=lambda c: (c.data or "").startswith("mod:nakladnoy:"))
    @guard
    def _click(call):
        ui.ack(bot, call)
        action = call.data.split(":", 2)[2]
        chat_id, tg_id = call.message.chat.id, call.from_user.id

        if action == "panel":
            _panel(bot, chat_id, tg_id)
        elif action == "yangi":
            sessions.set(tg_id, "nak:kutilmoqda", {})
            bot.send_message(
                chat_id,
                "Nakladnoy rasmini, PDF yoki Excel faylini yuboring.\n\n"
                "Rasm bo'lsa: hujjat to'liq ko'rinsin, matn o'qilarli bo'lsin.",
            )
        elif action.startswith("korish"):
            doc_id = int(action.split("_", 1)[1])
            _review(bot, chat_id, tg_id, doc_id)
        elif action.startswith("moslash"):
            doc_id = int(action.split("_", 1)[1])
            _match_screen(bot, chat_id, tg_id, doc_id)
        elif action.startswith("navbat"):
            doc_id = int(action.split("_", 1)[1])
            _next_pending(bot, chat_id, tg_id, doc_id)
        elif action.startswith("tanla"):
            _, item_id, product_id = action.split("_", 2)
            _pick(bot, chat_id, tg_id, int(item_id), product_id)
        elif action.startswith("yarat"):
            item_id = int(action.split("_", 1)[1])
            users.require_role(tg_id, "manager")
            note = bot.send_message(chat_id, "Bito'da yaratilmoqda…")
            created = create_in_bito(item_id)
            try:
                bot.delete_message(chat_id, note.message_id)
            except Exception:  # noqa: BLE001
                pass
            bot.send_message(
                chat_id,
                f"✅ Yaratildi va tanlandi: "
                f"<b>{ui.escape(created.get('name') or '')}</b>\n"
                f"SKU: <code>{ui.escape(str(created.get('sku') or ''))}</code>\n\n"
                "Sotish narxi va PLU (tarozi kodi) ni Bito'da belgilang — "
                "narx 0 bo'lib turibdi.",
                parse_mode="HTML")
            row = db.row("SELECT doc_id FROM nak_item WHERE tenant_id = ? "
                         "AND id = ?", (ctx.require(), item_id))
            _next_pending(bot, chat_id, tg_id, row["doc_id"])
        elif action.startswith("tashla"):
            item_id = int(action.split("_", 1)[1])
            skip_item(item_id)
            row = db.row("SELECT doc_id FROM nak_item WHERE tenant_id = ? "
                         "AND id = ?", (ctx.require(), item_id))
            _next_pending(bot, chat_id, tg_id, row["doc_id"])
        elif action.startswith("firma"):
            _, doc_id, supplier_id = action.split("_", 2)
            db.run("UPDATE nak_doc SET supplier_id = ? WHERE tenant_id = ? "
                   "AND id = ?", (supplier_id, ctx.require(), int(doc_id)))
            bot.send_message(chat_id, "Firma tanlandi.")
            _match_screen(bot, chat_id, tg_id, int(doc_id))
        elif action.startswith("yukla"):
            doc_id = int(action.split("_", 1)[1])
            _upload(bot, chat_id, tg_id, doc_id)
        elif action.startswith("bekor"):
            doc_id = int(action.split("_", 1)[1])
            db.run("UPDATE nak_doc SET status = 'bekor' WHERE tenant_id = ? "
                   "AND id = ?", (ctx.require(), doc_id))
            bot.send_message(chat_id, "Hujjat bekor qilindi.",
                             reply_markup=ui.main_menu(tg_id))

    @bot.message_handler(
        func=lambda m: sessions.get_global(m.from_user.id)[0] == "nak:kutilmoqda",
        content_types=["photo", "document"])
    @guard
    def _file(message):
        users.require_role(message.from_user.id, "manager")
        sessions.clear(message.from_user.id)
        _handle_file(bot, message)


def _handle_file(bot, message):
    tg_id, chat_id = message.from_user.id, message.chat.id

    if message.content_type == "photo":
        file_id = message.photo[-1].file_id
        filename, source = "nakladnoy.jpg", "photo"
    else:
        document = message.document
        file_id = document.file_id
        filename = document.file_name or "hujjat"
        if (document.file_size or 0) > MAX_FILE_MB * 1024 * 1024:
            bot.send_message(chat_id, f"Fayl {MAX_FILE_MB} MB dan katta.")
            return
        lowered = filename.lower()
        if lowered.endswith(".pdf"):
            source = "pdf"
        elif lowered.endswith((".xlsx", ".xls", ".xlsm")):
            source = "excel"
        elif lowered.endswith((".jpg", ".jpeg", ".png", ".webp")):
            source = "photo"
        else:
            bot.send_message(
                chat_id,
                "Bu turdagi fayl qo'llab-quvvatlanmaydi. "
                "Rasm, PDF yoki Excel yuboring.",
            )
            return

    doc_id = create_doc(tg_id, source, file_id)
    note = bot.send_message(chat_id, "Hujjat o'qilmoqda… bu 30–60 soniya "
                                     "olishi mumkin.")
    try:
        info = bot.get_file(file_id)
        data = bot.download_file(info.file_path)
        parsed = extract(data=data, filename=filename)
        save_doc(doc_id, parsed)
    except BotError as e:
        fail_doc(doc_id, e.user_message)
        bot.send_message(chat_id, f"⚠️ {e.user_message}")
        return
    except Exception as e:  # noqa: BLE001
        log.exception("Nakladnoy o'qishda xato")
        fail_doc(doc_id, str(e))
        bot.send_message(chat_id, "⚠️ Hujjatni o'qishda kutilmagan xato. "
                                  "Log yozildi.")
        return
    finally:
        try:
            bot.delete_message(chat_id, note.message_id)
        except Exception:  # noqa: BLE001
            pass

    _review(bot, chat_id, tg_id, doc_id)


def _review(bot, chat_id, tg_id, doc_id):
    doc = get_doc(doc_id)
    if not doc:
        bot.send_message(chat_id, "Hujjat topilmadi.")
        return
    items = get_items(doc_id)
    currency = tenant.get("currency_name") or "so'm"

    head = [f"<b>Nakladnoy</b> — {len(items)} ta qator"]
    if doc["supplier_name"]:
        head.append(f"Firma: {ui.escape(doc['supplier_name'])}")
    else:
        head.append("Firma: <i>aniqlanmadi</i>")
    if doc["doc_number"]:
        head.append(f"Hujjat №{ui.escape(doc['doc_number'])}")
    if doc["doc_date"]:
        head.append(f"Sana: {ui.escape(doc['doc_date'])}")

    lines = list(head) + [""]
    computed = 0.0
    for item in items:
        total = item["qty"] * item["block_size"] * item["price"]
        computed += total
        block = ""
        if item["block_size"] > 1:
            block = f" × {item['block_size']} dona"
        unit = f" {ui.escape(item['qty_unit'])}" if item["qty_unit"] else ""
        lines.append(
            f"• {ui.escape(item['raw_name'])}\n"
            f"    {item['qty']:g}{unit}{block} × {ui.money(item['price'])} = "
            f"{ui.money(total, currency)}"
        )

    lines += ["", f"<b>Jami: {ui.money(computed, currency)}</b>"]
    if doc["doc_total"]:
        diff = abs(computed - doc["doc_total"])
        percent = diff / doc["doc_total"] * 100 if doc["doc_total"] else 0
        if percent < 1:
            lines.append(f"✅ Hujjatdagi jami bilan mos: "
                         f"{ui.money(doc['doc_total'], currency)}")
        else:
            lines.append(
                f"⚠️ Hujjatda: {ui.money(doc['doc_total'], currency)} — "
                f"farq {percent:.1f}%. Miqdor yoki narx noto'g'ri o'qilgan "
                f"bo'lishi mumkin, qatorlarni tekshiring."
            )

    for chunk in ui.chunks("\n".join(lines)):
        bot.send_message(chat_id, chunk, parse_mode="HTML")

    bot.send_message(
        chat_id,
        "Ma'lumot to'g'ri bo'lsa — moslashtirishga o'tamiz.",
        reply_markup=ui.buttons(
            [("➡️ Moslashtirish", f"mod:nakladnoy:moslash_{doc_id}"),
             ("🗑 Bekor qilish", f"mod:nakladnoy:bekor_{doc_id}")],
            row_width=1, back="menu:root"),
    )


def _panel(bot, chat_id, tg_id):
    doc = last_doc(tg_id)
    lines = ["<b>Nakladnoy</b>", ""]
    if not ai.enabled():
        lines.append("⚠️ AI kaliti sozlanmagan — hujjat o'qish ishlamaydi. "
                     "Sotuvchiga murojaat qiling.")
    lines.append("Nakladnoy rasmini yuborasiz, bot uni o'qib chiqadi va "
                 "tekshirish uchun ko'rsatadi.")

    buttons = [("📄 Yangi hujjat", "mod:nakladnoy:yangi")]
    if doc and doc["status"] == "tekshirilmoqda":
        lines += ["", "Oxirgi hujjat: " + ui.escape(doc["supplier_name"] or "firma noma'lum")]
        buttons.insert(0, ("👁 Oxirgisini ko'rish",
                           f"mod:nakladnoy:korish_{doc['id']}"))
    elif doc and doc["status"] == "xato":
        lines += ["", f"⚠️ Oxirgi urinishda xato: {ui.escape(doc['error'] or '')}"]

    count = db.value(
        "SELECT COUNT(*) FROM nak_hint WHERE tenant_id = ?", (ctx.require(),),
        default=0)
    if count:
        lines += ["", f"O'rganilgan firma tuzilishi: {count} ta"]

    bot.send_message(chat_id, "\n".join(lines), parse_mode="HTML",
                     reply_markup=ui.buttons(buttons, back="menu:root"))


# ------------------------------------------------------- moslashtirish ekrani


def _match_screen(bot, chat_id, tg_id, doc_id):
    users.require_role(tg_id, "manager")
    doc = get_doc(doc_id)
    if not doc:
        bot.send_message(chat_id, "Hujjat topilmadi.")
        return

    if catalog.is_stale():
        bot.send_message(
            chat_id,
            "⚠️ Mahsulot katalogi eskirgan yoki bo'sh. "
            "Ombor → Yangilash tugmasini bosing, keyin qaytib keling.")
        return

    state = match_doc(doc_id) if doc["status"] == "tekshirilmoqda" \
        else summary(doc_id)

    lines = [
        "<b>Moslashtirish</b>", "",
        f"✅ Topildi: {len(state['matched'])} ta",
        f"❓ Qo'lda tanlash kerak: {len(state['pending'])} ta",
    ]
    if state["skipped"]:
        lines.append(f"⏭ Tashlab ketilgan: {len(state['skipped'])} ta")

    buttons = []
    if not doc["supplier_id"]:
        found, options = find_supplier(doc["supplier_name"])
        if found:
            db.run("UPDATE nak_doc SET supplier_id = ? WHERE tenant_id = ? "
                   "AND id = ?", (found["id"], ctx.require(), doc_id))
            lines += ["", f"Firma: {ui.escape(found.get('name') or '')}"]
        elif options:
            bot.send_message(
                chat_id,
                f"Firma «{ui.escape(doc['supplier_name'] or '')}» aniq "
                "topilmadi. Tanlang:",
                parse_mode="HTML",
                reply_markup=ui.buttons(
                    [(ui.escape(o.get("name") or "—"),
                      f"mod:nakladnoy:firma_{doc_id}_{o['id']}")
                     for o in options], row_width=1,
                    back=f"mod:nakladnoy:korish_{doc_id}"))
            return
        else:
            lines += ["", "⚠️ Firma topilmadi — Bito'da yetkazib beruvchi "
                          "yaratilgan bo'lishi kerak."]
    else:
        lines += ["", "Firma: tanlangan"]

    if state["pending"]:
        buttons.append(("❓ Qo'lda tanlash", f"mod:nakladnoy:navbat_{doc_id}"))
    if state["matched"] and doc["supplier_id"]:
        buttons.append((f"⬆️ Bito'ga yuklash ({len(state['matched'])} ta)",
                        f"mod:nakladnoy:yukla_{doc_id}"))
    buttons.append(("🗑 Bekor qilish", f"mod:nakladnoy:bekor_{doc_id}"))

    bot.send_message(chat_id, "\n".join(lines), parse_mode="HTML",
                     reply_markup=ui.buttons(buttons, row_width=1,
                                             back="menu:root"))


def _next_pending(bot, chat_id, tg_id, doc_id):
    state = summary(doc_id)
    if not state["pending"]:
        bot.send_message(chat_id, "Hamma qator moslashtirildi. 👍")
        _match_screen(bot, chat_id, tg_id, doc_id)
        return

    row = state["pending"][0]
    found = catalog.match(row["raw_name"], barcode=row["barcode"])
    left = len(state["pending"])

    lines = [f"<b>{ui.escape(row['raw_name'])}</b>",
             f"{row['qty']:g} × {ui.money(row['price'])}", "",
             f"Qolgan: {left} ta"]

    buttons = []
    if found["candidates"]:
        lines.append("")
        lines.append("Qaysi mahsulot?")
        for candidate in found["candidates"]:
            buttons.append((
                f"{candidate['name'][:45]}",
                f"mod:nakladnoy:tanla_{row['id']}_{candidate['product_id']}",
            ))
    else:
        lines.append("")
        lines.append("Katalogdan o'xshash mahsulot topilmadi.")
    buttons.append(("➕ Bito'da yangi yaratish",
                    f"mod:nakladnoy:yarat_{row['id']}"))
    buttons.append(("⏭ Tashlab ketish", f"mod:nakladnoy:tashla_{row['id']}"))

    bot.send_message(chat_id, "\n".join(lines), parse_mode="HTML",
                     reply_markup=ui.buttons(
                         buttons, row_width=1,
                         back=f"mod:nakladnoy:moslash_{doc_id}"))


def _pick(bot, chat_id, tg_id, item_id, product_id):
    row = db.row("SELECT * FROM nak_item WHERE tenant_id = ? AND id = ?",
                 (ctx.require(), item_id))
    if not row:
        return
    match = db.row("SELECT name FROM catalog WHERE tenant_id = ? "
                   "AND product_id = ?", (ctx.require(), product_id))
    set_match(item_id, product_id, match["name"] if match else None)
    bot.send_message(chat_id, "Eslab qoldim — keyingi nakladnoylarda "
                              "avtomatik qo'llanadi.")
    _next_pending(bot, chat_id, tg_id, row["doc_id"])


def _upload(bot, chat_id, tg_id, doc_id):
    users.require_role(tg_id, "manager")
    doc = get_doc(doc_id)
    state = summary(doc_id)
    if not doc["supplier_id"]:
        bot.send_message(chat_id, "Avval firmani tanlang.")
        return
    if not state["matched"]:
        bot.send_message(chat_id, "Moslashtirilgan qator yo'q.")
        return

    note = bot.send_message(chat_id, "Bito'ga yuklanmoqda…")
    items = [dict(row) for row in state["matched"]]
    result = nak_upload.upload(items, doc["supplier_id"])
    nak_upload.mark_uploaded(doc_id, result)

    try:
        bot.delete_message(chat_id, note.message_id)
    except Exception:  # noqa: BLE001
        pass

    currency = tenant.get("currency_name") or "so'm"
    lines = []
    if result["numbers"]:
        lines.append("✅ Bito'ga yuklandi")
        lines.append(f"Kirim raqami: {', '.join(result['numbers'])}")
        lines.append(f"{result['uploaded_count']} ta mahsulot, "
                     f"{ui.money(result['uploaded_total'], currency)}")
    if result["failed"]:
        lines.append("")
        lines.append(f"⚠️ {len(result['failed'])} ta partiya yuklanmadi:")
        for fail in result["failed"]:
            lines.append(f"• {fail['count']} ta mahsulot: "
                         f"{ui.escape(fail['error'])}")
        lines.append("")
        lines.append("Yuklanmagan qismini qayta yuklash uchun nakladnoyni "
                     "qaytadan yuboring.")
    if state["skipped"]:
        lines.append(f"\n⏭ {len(state['skipped'])} ta qator tashlab ketildi.")

    for chunk in ui.chunks("\n".join(lines)):
        bot.send_message(chat_id, chunk, parse_mode="HTML")
    bot.send_message(chat_id, "—", reply_markup=ui.main_menu(tg_id))
